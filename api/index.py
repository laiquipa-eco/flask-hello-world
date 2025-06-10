import os
from openpyxl import load_workbook
import io
from flask import jsonify, Response


def handler(request):
    # Obtiene los datos del JSON
    data = request.get_json()

    if not data or not isinstance(data, dict):
        return jsonify(
            {"error": "Se requiere un JSON con las celdas y los valores a modificar"}
        ), 400

    try:
        # Cargar la plantilla de Excel
        wb = load_workbook("plantilla.xlsx")
        ws = wb.active  # Seleccionamos la primera hoja

        # Recorrer las celdas en el JSON y actualizar los valores en el archivo
        for celda, valor in data.items():
            try:
                # Intentar actualizar la celda. No necesitamos verificar si la celda existe.
                ws[celda.upper()] = valor
            except KeyError:
                # Si ocurre un KeyError, significa que la celda no es válida.
                return jsonify(
                    {"error": f"La celda {celda} no existe en la plantilla."}
                ), 400

        # Guardamos el archivo modificado en memoria (en un objeto BytesIO)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Creamos una respuesta con el archivo en memoria
        response = Response(
            output.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["Content-Disposition"] = (
            "attachment; filename=archivo_modificado.xlsx"
        )

        # Retornamos la respuesta
        return response
    except Exception as e:
        return jsonify({"error": str(e)}), 500
